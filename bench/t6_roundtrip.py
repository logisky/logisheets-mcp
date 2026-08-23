"""T6 — hand back a file you were given, with the parts you don't model intact.

A workbook that arrives from a person carries things no agent tool has a concept
of: conditional formatting, a table, page setup for printing, frozen panes,
data validation, a defined name, a merged title. Writing one cell should not
cost any of it.

The fixture is built with openpyxl — neutral, belongs to none of the entrants —
because it can author all of those features. Each server then opens the file,
writes ONE cell, and saves; the parts are compared before and after.

Scored per part: kept (byte-identical), changed (present, differs), or lost.
"Changed" is not automatically a failure — a writer may legitimately reorder
attributes — so the specific features are also checked by looking for their
elements.
"""
import os
import shutil
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mcp_client import Server                      # noqa: E402
from contestants import CONTESTANTS, WORK          # noqa: E402

SEED = os.path.join(WORK, "t6-rich.xlsx")

# The features to look for afterwards: a label, and something in the XML that
# only exists if the feature survived.
FEATURES = {
    "conditional formatting": "conditionalFormatting",
    "table / ListObject": "tableParts",
    "page setup": "pageSetup",
    "frozen panes": "pane",
    "data validation": "dataValidation",
    "merged cells": "mergeCell",
    "defined name": "definedName",
    "column widths": "<col ",
}


def build_seed() -> None:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Quarterly report"
    ws.merge_cells("A1:D1")
    for j, h in enumerate(["region", "q1", "q2", "total"], start=1):
        ws.cell(row=2, column=j, value=h)
    for i, (region, q1, q2) in enumerate(
        [("north", 10, 12), ("south", 20, 18), ("east", 30, 35)], start=3
    ):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=q1)
        ws.cell(row=i, column=3, value=q2)
        ws.cell(row=i, column=4, value=f"=B{i}+C{i}")
    # A table over the same region.
    table = Table(displayName="Sales", ref="A2:D5")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True
    )
    ws.add_table(table)
    # Conditional formatting: highlight totals over 40.
    ws.conditional_formatting.add(
        "D3:D5",
        CellIsRule(operator="greaterThan", formula=["40"],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE")),
    )
    # Print setup, frozen header, a validation rule, a name, a column width.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.freeze_panes = "A3"
    dv = DataValidation(type="whole", operator="between", formula1="0",
                        formula2="1000", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B3:C5")
    ws.column_dimensions["A"].width = 18
    wb.defined_names.add(DefinedName("HeaderRow", attr_text="Report!$A$2:$D$2"))
    wb.save(SEED)


def parts(path: str) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as z:
        return {n: z.read(n) for n in z.namelist() if not n.endswith("/")}


def features_present(path: str) -> dict[str, bool]:
    blob = b"".join(parts(path).values())
    return {
        label: token.encode("utf8") in blob for label, token in FEATURES.items()
    }


def run_logisheets(path: str) -> None:
    c = CONTESTANTS["logisheets"]
    with Server("logisheets", c["command"], c["env"]) as s:
        s.call("open_workbook", {"path": path})
        s.call("set_cells", {"sheetIdx": 0,
                             "cells": [{"row": 6, "col": 0, "content": "note"}]})
        s.call("save_workbook", {"path": path})


def run_excel_mcp(path: str) -> None:
    c = CONTESTANTS["excel-mcp-server"]
    with Server("excel-mcp-server", c["command"], c["env"]) as s:
        s.call("write_data_to_excel", {"filepath": path, "sheet_name": "Report",
                                       "data": [["note"]], "start_cell": "A7"})


def run_spreadsheet_kit(path: str) -> None:
    c = CONTESTANTS["spreadsheet-kit"]
    with Server("spreadsheet-kit", c["command"], c["env"]) as s:
        wbs, _ = s.call("list_workbooks", {"include_paths": True})
        wid = next((w["workbook_id"] for w in (wbs or {}).get("workbooks", [])
                    if w.get("path") == os.path.basename(path)), None)
        if wid is None:
            raise RuntimeError("not in workspace")
        fork, _ = s.call("create_fork", {"workbook_or_fork_id": wid})
        fid = fork["fork_id"]
        s.call("edit_batch", {"fork_id": fid, "sheet_name": "Report",
                              "edits": [{"address": "A7", "value": "note"}]})
        # `target_path` defaults to the original, but say it outright: a fork
        # that is never written back would score a perfect round trip for
        # having done nothing.
        # It declines to overwrite the original without a flag, which is a
        # reasonable default — so it writes a sibling and that is what gets
        # compared. Nothing about the result turns on the filename.
        out = os.path.basename(path).replace(".xlsx", "-saved.xlsx")
        r, raw = s.call("save_fork", {"fork_id": fid, "target_path": out})
        if r is None:
            raise RuntimeError(f"save_fork: {raw[:200]}")
        saved = os.path.join(os.path.dirname(path), out)
        if not os.path.exists(saved):
            raise RuntimeError(f"save_fork reported ok but wrote nothing: {raw[:160]}")
        shutil.copyfile(saved, path)


RUNNERS = {
    "logisheets": run_logisheets,
    "spreadsheet-kit": run_spreadsheet_kit,
    "excel-mcp-server": run_excel_mcp,
}

if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    build_seed()
    before_parts = parts(SEED)
    before_feats = features_present(SEED)
    missing = [k for k, v in before_feats.items() if not v]
    print("T6 — write one cell into a file you were handed, keep the rest.\n")
    print(f"    fixture: {len(before_parts)} parts, features present: "
          f"{sum(before_feats.values())}/{len(FEATURES)}"
          + (f" (fixture lacks {missing})" if missing else ""))
    print()

    for name, runner in RUNNERS.items():
        path = os.path.join(WORK, f"t6-{name}.xlsx")
        shutil.copyfile(SEED, path)
        try:
            runner(path)
        except Exception as e:
            print(f"--- {name} ---\n    failed: {type(e).__name__}: {e}\n")
            continue
        after = parts(path)
        kept = sum(1 for n, b in before_parts.items() if after.get(n) == b)
        changed = sum(1 for n, b in before_parts.items()
                      if n in after and after[n] != b)
        lost = [n for n in before_parts if n not in after]
        feats = features_present(path)
        gone = [k for k, v in before_feats.items() if v and not feats[k]]
        wrote = any(b"note" in b for b in after.values())
        print(f"--- {name} ---")
        print(f"    the edit landed : {'yes' if wrote else 'NO'}")
        print(f"    parts           : {kept} identical, {changed} changed, "
              f"{len(lost)} lost")
        if lost:
            print(f"      lost: {sorted(lost)}")
        print(f"    features kept   : "
              f"{sum(1 for k, v in before_feats.items() if v and feats[k])}"
              f"/{sum(before_feats.values())}")
        if gone:
            print(f"      \033[31mlost: {gone}\033[0m")
        print()
