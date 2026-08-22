"""T1 — does a formula written in this session return a value?

The seed workbook is built with openpyxl: a third-party library belonging to
none of the contestants, so nobody starts from a file their own writer
produced. Every server is then driven with its own idioms, taken from its own
tool descriptions, and given its own recalculation step if it offers one.

Expected answer: 30, from 10 + 20. Derived here, not from any server.
"""
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mcp_client import Server                      # noqa: E402
from contestants import CONTESTANTS, WORK          # noqa: E402

EXPECTED = 30
SEED = os.path.join(WORK, "seed.xlsx")


def build_seed() -> None:
    """A1=10, A2=20, nothing else. Written by openpyxl on purpose."""
    sys.path.insert(0, os.path.join(WORK, ".venv", "lib"))
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["A2"] = 10, 20
    wb.save(SEED)


def fresh_copy(tag: str) -> str:
    path = os.path.join(WORK, f"t1-{tag}.xlsx")
    shutil.copyfile(SEED, path)
    return path


def run_logisheets(s: Server, path: str):
    s.call("open_workbook", {"path": path})
    s.call("set_cells", {"sheetIdx": 0, "cells": [
        {"row": 2, "col": 0, "content": "=SUM(A1:A2)"}]})
    got, raw = s.call("get_cells", {"sheetIdx": 0, "startRow": 2, "startCol": 0,
                                    "endRow": 2, "endCol": 0})
    cells = (got or {}).get("cells", [])
    return (cells[0].get("value") if cells else None), raw


def run_excel_mcp(s: Server, path: str):
    _, raw1 = s.call("apply_formula", {
        "filepath": path, "sheet_name": "Sheet", "cell": "A3",
        "formula": "=SUM(A1:A2)"})
    got, raw = s.call("read_data_from_excel", {
        "filepath": path, "sheet_name": "Sheet",
        "start_cell": "A3", "end_cell": "A3"})
    return got, f"{raw1}\n---\n{raw}"


def run_spreadsheet_kit(s: Server, path: str):
    # The workspace is scanned at startup; a workbook is addressed by the
    # `workbook_id` this listing hands back, not by its filename.
    wbs, rawl = s.call("list_workbooks", {"include_paths": True})
    wid = None
    for w in (wbs or {}).get("workbooks", []) if isinstance(wbs, dict) else []:
        if w.get("path") == os.path.basename(path):
            wid = w.get("workbook_id")
    if wid is None:
        return None, f"no workbook_id for {os.path.basename(path)}: {rawl}"
    fork, rawf = s.call("create_fork", {"workbook_or_fork_id": wid})
    fid = (fork or {}).get("fork_id") if isinstance(fork, dict) else None
    if fid is None:
        return None, f"create_fork gave no fork_id: {rawf}"
    _, rawe = s.call("edit_batch", {
        "fork_id": fid, "sheet_name": "Sheet",
        "edits": [{"address": "A3", "formula": "=SUM(A1:A2)"}]})
    _, rawr = s.call("recalculate", {"fork_id": fid})
    got, raw = s.call("inspect_cells", {
        "workbook_or_fork_id": fid, "sheet_name": "Sheet", "targets": ["A3"]})
    return got, f"{rawe}\n---\n{rawr}\n---\n{raw}"


RUNNERS = {
    "logisheets": run_logisheets,
    "excel-mcp-server": run_excel_mcp,
    "spreadsheet-kit": run_spreadsheet_kit,
}

if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    build_seed()
    print(f"T1 — write =SUM(A1:A2) over 10 and 20, read it back. Expect {EXPECTED}.\n")
    for name, runner in RUNNERS.items():
        c = CONTESTANTS[name]
        path = fresh_copy(name)
        try:
            with Server(name, c["command"], c["env"]) as s:
                value, raw = runner(s, path)
                print(f"--- {name} ({s.calls} tool calls) ---")
                print(f"    returned: {str(value)[:300]}")
                print(f"    raw     : {' '.join(str(raw).split())[:300]}")
        except Exception as e:
            print(f"--- {name} ---\n    failed: {type(e).__name__}: {e}")
        print()
