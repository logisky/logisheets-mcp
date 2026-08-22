"""T12 — the same comprehension questions, on a file nobody's server built.

Also added after the pre-registered list. T11 let every server read back its own
output, which is the flattering version: ours had blocks in it because we put
them there. The obvious objection is that our advantage exists only on files we
authored, and this is the test of that objection.

The workbook here is a plain A1 spreadsheet with a header row and labels in
column A — what a person hands you — written by openpyxl so it belongs to none
of the contestants. No blocks, no names, no schema.

Same three questions as T11. The measurement is which part of the advantage
survives adoption, and at what cost.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mcp_client import Server                      # noqa: E402
from contestants import CONTESTANTS, WORK          # noqa: E402
from t11_coldread import names_or_coordinates      # noqa: E402

A = dict(base=1000.0, growth=0.08, margin=0.20, tax=0.25,
         wacc=0.10, tg=0.025, net_debt=500.0, shares=100.0)
ORDER = ["base", "growth", "margin", "tax", "wacc", "tg", "net_debt", "shares"]
EXPECTED_PER_SHARE = 20.803603425995494
SEED = os.path.join(WORK, "t12-human.xlsx")


def build_human_file() -> None:
    """A DCF as a person would leave it: labels down column A, a header row
    over the projection, formulas in coordinates. Header row 10, data 11..15."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    for i, k in enumerate(ORDER, start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=A[k])
    for j, h in enumerate(["year", "revenue", "fcf", "discount_factor",
                           "present_value"], start=1):
        ws.cell(row=10, column=j, value=h)
    for r, n in zip(range(11, 16), (1, 2, 3, 4, 5)):
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=f"=$B$1*POWER(1+$B$2,A{r})")
        ws.cell(row=r, column=3, value=f"=B{r}*$B$3*(1-$B$4)")
        ws.cell(row=r, column=4, value=f"=1/POWER(1+$B$5,A{r})")
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
    for r, (label, formula) in enumerate([
        ("pv_explicit", "=SUM(E11:E15)"),
        ("pv_terminal", "=C15*(1+$B$6)/($B$5-$B$6)*D15"),
        ("enterprise_value", "=B17+B18"),
        ("equity", "=B19-$B$7"),
        ("per_share", "=B20/$B$8"),
    ], start=17):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=formula)
    wb.save(SEED)


def run_logisheets(path: str):
    c = CONTESTANTS["logisheets"]
    out = {}
    with Server("logisheets", c["command"], c["env"]) as s:
        s.call("open_workbook", {"path": path})
        blocks, _ = s.call("list_blocks")
        groups = blocks if isinstance(blocks, list) else []
        out["blocks_found"] = sum(len(g.get("blocks", [])) for g in groups)
        # Sanity: the file computes, so the model was read correctly.
        cells, _ = s.call("get_cells", {"sheetIdx": 0, "startRow": 20,
                                        "startCol": 1, "endRow": 20,
                                        "endCol": 1})
        cs = (cells or {}).get("cells", [])
        out["per_share"] = cs[0].get("value") if cs else None
        out["calls_before_adopt"] = s.calls
        out["bytes_before_adopt"] = s.bytes
        # Adopt the projection: field names come from the header row.
        r, raw = s.call("convert_to_block", {
            "sheet": "Model", "name": "proj",
            "position": {"row": 10, "col": 0},
            "row_count": 5, "col_count": 5, "header_row": 9})
        out["convert"] = ("ok" if r is not None else f"ERR {raw[:160]}")
        d, _ = s.call("describe_block", {"name": "proj"})
        out["fields"] = [f["name"] for f in (d or {}).get("fields", [])]
        out["keys"] = (d or {}).get("keys")
        out["q2_rule"] = next((f.get("value_formula")
                               for f in (d or {}).get("fields", [])
                               if f["name"] == "fcf"), None)
        # What a cell of the adopted column actually holds now.
        got, _ = s.call("get_cells", {"sheetIdx": 0, "startRow": 10,
                                      "startCol": 2, "endRow": 10, "endCol": 2})
        cc = (got or {}).get("cells", [])
        out["q2_cell_formula"] = cc[0].get("formula") if cc else None
        # Does addressing by name work after adoption?
        ev, raw_ev = s.call("eval_formula",
                            {"expr": '=SUM(BLOCKREFS("proj","*","fcf"))'})
        out["blockrefs"] = ev if ev is not None else raw_ev[:120]
        out["calls"], out["bytes"] = s.calls, s.bytes
    return out


def run_other(name: str, path: str, sheet: str):
    c = CONTESTANTS[name]
    out = {}
    with Server(name, c["command"], c["env"]) as s:
        if name == "spreadsheet-kit":
            wbs, _ = s.call("list_workbooks", {"include_paths": True})
            wid = next((w["workbook_id"] for w in (wbs or {}).get("workbooks", [])
                        if w.get("path") == os.path.basename(path)), None)
            if wid is None:
                return {"error": "not in workspace"}
            ov, _ = s.call("sheet_overview", {"workbook_or_fork_id": wid,
                                              "sheet_name": sheet})
            regs = (ov or {}).get("detected_regions", []) if isinstance(ov, dict) else []
            out["regions"] = [(r.get("bounds"), r.get("region_kind")) for r in regs]
            ins, _ = s.call("inspect_cells", {"workbook_or_fork_id": wid,
                                              "sheet_name": sheet,
                                              "targets": ["C11"]})
            cells = (ins or {}).get("cells", []) if isinstance(ins, dict) else []
            out["q2_rule"] = cells[0].get("formula") if cells else None
        else:
            got, _ = s.call("read_data_from_excel", {
                "filepath": path, "sheet_name": sheet,
                "start_cell": "A1", "end_cell": "E22"})
            cells = []
            if isinstance(got, dict) and isinstance(got.get("result"), str):
                import json as _json
                try:
                    cells = _json.loads(got["result"]).get("cells", [])
                except Exception:
                    pass
            by = {x["address"]: x.get("value") for x in cells}
            out["q2_rule"] = by.get("C11")
        out["calls"], out["bytes"] = s.calls, s.bytes
    return out


if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    build_human_file()
    print("T12 — a plain A1 workbook written by openpyxl: no blocks, no names.\n"
          "      Does adoption transfer the advantage? "
          "(added after the pre-registered list)\n")

    r = run_logisheets(SEED)
    print("--- logisheets ---")
    print(f"    blocks in the file as handed over : {r['blocks_found']}")
    ps = r["per_share"]
    print(f"    per_share reads                   : "
          f"{ps if not isinstance(ps, float) else round(ps, 6)} "
          f"{'✓' if isinstance(ps, float) and abs(ps - EXPECTED_PER_SHARE) < 1e-6 else '✗'}")
    print(f"    convert_to_block                  : {r['convert']}")
    print(f"    field names adopted               : {r['fields']}")
    print(f"    row keys adopted                  : {r['keys']}")
    print(f"    BLOCKREFS by name now works       : {r['blockrefs']}")
    print(f"    Q2 as a field RULE                : {r['q2_rule']}")
    print(f"    Q2 as the cell's own formula      : {r['q2_cell_formula']}")
    print(f"                                        -> refers by "
          f"{names_or_coordinates(r['q2_cell_formula'])}")
    print(f"    cost  : {r['calls_before_adopt']} calls / "
          f"{r['bytes_before_adopt']} bytes to orient, then "
          f"{r['calls'] - r['calls_before_adopt']} more to adopt and read back "
          f"({r['bytes'] - r['bytes_before_adopt']} bytes)\n")

    for name in ("spreadsheet-kit", "excel-mcp-server"):
        o = run_other(name, SEED, "Model")
        print(f"--- {name} ---")
        if "error" in o:
            print(f"    {o['error']}\n")
            continue
        if "regions" in o:
            print(f"    regions detected : {o['regions']}")
        print(f"    Q2 fcf rule      : {o['q2_rule']}")
        print(f"                       -> refers by "
              f"{names_or_coordinates(o['q2_rule'])}")
        print(f"    cost             : {o['calls']} calls, {o['bytes']} bytes\n")
