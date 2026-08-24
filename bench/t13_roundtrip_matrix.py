"""T13 — a per-feature round-trip matrix for this engine alone.

T6 asked whether a handed-over file survives one edit, against other servers.
This asks a narrower question in more detail: of the things real workbooks
actually contain, which ones come back? No competitors — the output is a punch
list.

Each feature is authored by openpyxl, then looked for in the saved output by a
token that only exists if it survived. A token is crude but it is checkable: it
does not depend on our own reader agreeing with our own writer.
"""
import os
import shutil
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mcp_client import Server                      # noqa: E402
from contestants import CONTESTANTS, WORK          # noqa: E402

SEED = os.path.join(WORK, "t13-everything.xlsx")

# (label, token, which part it should be in) — '*' means anywhere in the package.
FEATURES = [
    ("number format: date", "yyyy-mm-dd", "*"),
    ("number format: percent", "0.00%", "*"),
    # Written as `&quot;$&quot;#,##0.00`; match the part that survives escaping.
    ("number format: currency", "#,##0.00", "*"),
    # We write `<b/>`, openpyxl writes `<b val="1"/>`; match either.
    ("bold font", "<b", "xl/styles.xml"),
    ("fill colour", "FFFF00", "xl/styles.xml"),
    ("cell border", "<border", "xl/styles.xml"),
    ("comment / note", "This needs review", "*"),
    ("hyperlink", "hyperlink", "xl/worksheets/sheet1.xml"),
    ("hidden row", 'hidden="1"', "xl/worksheets/sheet1.xml"),
    ("column outline level", "outlineLevel", "xl/worksheets/sheet1.xml"),
    ("frozen panes", "<pane ", "xl/worksheets/sheet1.xml"),
    ("merged cells", "mergeCell", "xl/worksheets/sheet1.xml"),
    ("data validation", "dataValidation", "xl/worksheets/sheet1.xml"),
    ("conditional formatting", "conditionalFormatting", "xl/worksheets/sheet1.xml"),
    ("page setup", "pageSetup", "xl/worksheets/sheet1.xml"),
    ("autofilter", "autoFilter", "*"),
    ("defined name", "MyRange", "xl/workbook.xml"),
    ("second sheet kept", "Notes", "xl/workbook.xml"),
    ("hidden sheet", 'state="hidden"', "xl/workbook.xml"),
    ("tab colour", "tabColor", "*"),
    ("docProps: creator", "Jane Analyst", "docProps/core.xml"),
    ("docProps: title", "Q3 numbers", "docProps/core.xml"),
    # Known lost, with a diagnosis rather than a shrug — see the note below.
    ("chart", "chart", "*"),
]

# Why the chart is lost, as of writing:
#
#   1. `xl/drawings/drawingN.xml` is matched by literal prefixed element name
#      (`xdr:wsDr`, `xdr:twoCellAnchor`). openpyxl declares the
#      spreadsheetDrawing namespace as the DEFAULT one, so its elements carry no
#      prefix and no anchor is seen at all — the chart is skipped in silence.
#      Same failure class as keying on `t="inlineStr"`: one valid serialization
#      recognised out of two. Images travel this code path too.
#   2. Only `twoCellAnchor` is modelled. openpyxl's default is `oneCellAnchor`
#      (a `from` plus an `ext` size), so even with the prefix handled its charts
#      would still be skipped.
#
# Both were confirmed by hand: a chart written with an explicit twoCellAnchor is
# dropped as well, which rules the anchor kind out as the only cause.


def build_seed() -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    wb.properties.creator = "Jane Analyst"
    wb.properties.title = "Q3 numbers"

    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Quarterly report"
    ws.merge_cells("A1:D1")
    for j, h in enumerate(["region", "amount", "share", "due"], start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin = Side(style="thin")
    import datetime
    for i, (region, amount, share) in enumerate(
        [("north", 1234.5, 0.25), ("south", 2345.75, 0.35), ("east", 3456.25, 0.4)],
        start=3,
    ):
        ws.cell(row=i, column=1, value=region).border = Border(
            left=thin, right=thin, top=thin, bottom=thin
        )
        a = ws.cell(row=i, column=2, value=amount)
        a.number_format = '"$"#,##0.00'
        s = ws.cell(row=i, column=3, value=share)
        s.number_format = "0.00%"
        d = ws.cell(row=i, column=4, value=datetime.date(2026, 3, i))
        d.number_format = "yyyy-mm-dd"
    ws["A3"].comment = Comment("This needs review", "Jane")
    ws["A7"] = "see the docs"
    ws["A7"].hyperlink = "https://example.com/docs"
    ws.row_dimensions[6].hidden = True
    ws.column_dimensions["D"].outline_level = 1
    ws.freeze_panes = "A3"
    dv = DataValidation(type="list", formula1='"north,south,east"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A3:A5")
    ws.conditional_formatting.add(
        "B3:B5",
        CellIsRule(operator="greaterThan", formula=["2000"],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE")),
    )
    ws.page_setup.orientation = "landscape"
    ws.auto_filter.ref = "A2:D5"
    ws.sheet_properties.tabColor = "1072BA"

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=2, max_row=5), titles_from_data=True)
    ws.add_chart(chart, "F3")

    notes = wb.create_sheet("Notes")
    notes["A1"] = "kept for reference"
    hidden = wb.create_sheet("Hidden")
    hidden["A1"] = "not for display"
    hidden.sheet_state = "hidden"

    wb.defined_names.add(DefinedName("MyRange", attr_text="Data!$A$2:$D$5"))
    wb.save(SEED)


def present(path: str) -> dict[str, bool]:
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        blobs = {n: z.read(n) for n in names if not n.endswith("/")}
    whole = b"".join(blobs.values())
    out = {}
    for label, token, where in FEATURES:
        t = token.encode("utf8")
        if where == "*":
            out[label] = t in whole
        else:
            # A part may be renamed (sheet1 vs sheet2); fall back to the package.
            hit = any(t in b for n, b in blobs.items() if n == where)
            if not hit and where.startswith("xl/worksheets/"):
                hit = any(
                    t in b for n, b in blobs.items() if n.startswith("xl/worksheets/")
                )
            out[label] = hit
    return out


if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    build_seed()
    before = present(SEED)
    absent = [k for k, v in before.items() if not v]
    if absent:
        print(f"NOTE: the fixture itself lacks {absent} — not scored\n")

    path = os.path.join(WORK, "t13-saved.xlsx")
    shutil.copyfile(SEED, path)
    c = CONTESTANTS["logisheets"]
    with Server("logisheets", c["command"], c["env"]) as s:
        r, raw = s.call("open_workbook", {"path": path})
        opened = r is not None and "unreachable" not in str(raw)
        print(f"open: {'ok' if opened else 'FAILED — ' + str(raw)[:120]}")
        s.call("set_cells", {"sheetIdx": 0,
                             "cells": [{"row": 9, "col": 0, "content": "touched"}]})
        s.call("save_workbook", {"path": path})
    after = present(path)

    kept = [k for k, v in before.items() if v and after[k]]
    lost = [k for k, v in before.items() if v and not after[k]]
    scored = len(kept) + len(lost)
    print(f"\nround-trip: {len(kept)}/{scored} features kept\n")
    for label, _, _ in FEATURES:
        if not before[label]:
            continue
        print(f"  {'✓' if after[label] else '✗'} {label}")
    if lost:
        print(f"\nlost: {lost}")
